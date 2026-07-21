#!/usr/bin/env python3
import http.server, json, os, uuid, base64, mimetypes, requests, subprocess, time, random, threading, sys, io, tempfile, queue as _queue_mod
sys.stdout.reconfigure(line_buffering=True)  # noqa
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from concurrent.futures import ThreadPoolExecutor

LLAMA_BASE = "http://localhost:8081"
LLAMA_URL = f"{LLAMA_BASE}/v1/chat/completions"

SEARXNG_URL = "http://localhost:8080/search"
COMFYUI_URL = "http://localhost:8188"
HOST, PORT = "0.0.0.0", 3000

with open(os.path.expanduser("~/local-ai-files/model.txt"), "r") as file:
    MODEL_ID = file.read()

import sys
COMFYUI_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ComfyUI")
sys.path.insert(0, COMFYUI_DIR)
COMFYUI_OUTPUT = os.path.expanduser("~/local-ai-files/ComfyUI/output")
LLAMA_SERVER_PATH = os.path.expanduser("~/local-ai/llama.cpp/build/bin/llama-server")
LLAMA_SERVER_ARGS = ["--host", "0.0.0.0", "--port", "8081", "--models-dir", os.path.expanduser("~/local-ai-files/my-models/"), "--n-gpu-layers", "12", "--no-kv-offload", "--cache-type-k", "q8_0", "--cache-type-v", "q8_0", "--ctx-size", "16384"]
SESSIONS_FILE = os.path.expanduser("~/local-ai-files/sessions.json")
IMG_PATH = os.path.expanduser("~/local-ai-files/ComfyUI/output")
PROMPT_PATH = os.path.expanduser("~/local-ai-files/sys_prompt.txt")

IMAGE_MODELS = {
    "sd3_5_medium": {
        "unet": "sd3.5_medium-Q4_K_M.gguf",
        "clip1": "clip_l.safetensors",
        "clip2": "clip_g.safetensors",
        "t5": "t5-v1_1-xxl-encoder-Q4_K_M.gguf",
        "vae": "sd3_vae.safetensors",
        "description": "High-quality Stable Diffusion 3.5 Medium. Can edit and draw or generate any kind of photo. High Quality but Slow.",
    },
    "realistic": {
        "ckpt": "Realistic_Vision_V6.0_NV_B1_fp16.safetensors",
        "vae": "vae-ft-mse-840000-ema-pruned.safetensors",
        "description": "Hyper-realistic photos and lifelike images. Moderately fast but not the best quality.",
    },
    "sketch": {
        "ckpt": "dreamshaper_8.safetensors",
        "vae": "vae-ft-mse-840000-ema-pruned.safetensors",
        "description": "Pencil sketches, line art, and artistic drawings (prompt with 'pencil sketch'). Fast and good.",
    },
    "ghibli": {
        "ckpt": "ghibli_diffusion_v1.ckpt",
        "vae": "vae-ft-mse-840000-ema-pruned.safetensors",
        "description": "Studio Ghibli style anime illustrations (prompt with 'ghibli style'). Fast but not the best.",
    },
}

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": "Search the web for real-time/current information. Use this for weather, news, sports, stock prices, recent events, or any query where up-to-date data matters. Do NOT answer time-sensitive questions from memory — always search.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "The search query"}
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "generate_image",
            "description": "Generate or draw an image. You MUST choose a style model.",
            "parameters": {
                "type": "object",
                "properties": {
                    "prompt": {
                        "type": "string",
                        "description": "Detailed visual description of what to draw/generate.",
                    },
                    "negative_prompt": {
                        "type": "string",
                        "description": "Things to avoid in the image",
                    },
                    "model": {
                        "type": "string",
                        "enum": list(IMAGE_MODELS.keys()),
                        "description": "Art style to use. Options: "
                        + ", ".join(
                            [
                                f"'{k}' ({v['description']})"
                                for k, v in IMAGE_MODELS.items()
                            ]
                        ),
                    },
                },
                "required": ["prompt", "model"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "edit_image",
            "description": "Edit or modify an existing image (Img2Img). Use this when the user asks to change, restyle, recolor, add to, or modify a previously generated image OR an uploaded image.",
            "parameters": {
                "type": "object",
                "properties": {
                    "prompt": {
                        "type": "string",
                        "description": "Detailed visual description of what the edited image should look like.",
                    },
                    "negative_prompt": {
                        "type": "string",
                        "description": "Things to avoid in the edited image.",
                    },
                    "denoise": {
                        "type": "number",
                        "description": "Denoising strength between 0.1 and 1.0. Use 0.3-0.5 for subtle edits/recoloring, 0.5-0.7 for strong edits maintaining structure, and 0.7-0.9 for heavy transformations.",
                    },
                },
                "required": ["prompt"],
            },
        },
    },
]


with open(PROMPT_PATH, "r") as file:
    SYS_CONTENT = file.read()
model_list = "; ".join(f"{k}: {v['description']}" for k, v in IMAGE_MODELS.items())
SYS_CONTENT = SYS_CONTENT.replace("%model_list%", model_list)
SYS_CONTENT = SYS_CONTENT.replace("%_image_keys%", str(list(IMAGE_MODELS.keys())))

print("Prompt:\n", "*" * 80, "\n", SYS_CONTENT, "\n", "*" * 80)

sessions = {}
sessions_meta = {}
tasks = {}
model_status = "unloaded"
_last_tps = None
_last_llm_use = time.time()

_data_lock = threading.Lock()

MAX_QUEUE_SIZE = 5
_task_queue = []
_queue_lock = threading.Lock()
_queue_cond = threading.Condition(_queue_lock)
_current_task_id = None

MAX_INPUT_TOKENS = 3000

_event_queue = _queue_mod.Queue()
_llm_pool = ThreadPoolExecutor(max_workers=1)
_tool_pool = ThreadPoolExecutor(max_workers=2)

_overheated = False
_gpu_temp = None
TEMP_THRESHOLD_ON = 85
TEMP_THRESHOLD_OFF = 65
RAM_EVAC_THRESHOLD = 95
RAM_RESUME_THRESHOLD = 50
_ram_evacuating = False


def load_sessions():
    global sessions, sessions_meta
    try:
        with open(SESSIONS_FILE) as f:
            data = json.load(f)
        with _data_lock:
            sessions = {}
            sessions_meta = {}
            for sid, sdata in data.get("sessions", {}).items():
                sessions[sid] = sdata.get("messages", [])
                sessions_meta[sid] = {
                    "name": sdata.get("name", "Chat"),
                    "created": sdata.get("created", time.time()),
                    "updated": sdata.get("updated", time.time()),
                }
    except (FileNotFoundError, json.JSONDecodeError):
        with _data_lock:
            sessions = {}
            sessions_meta = {}


def save_sessions():
    with _data_lock:
        data = {"sessions": {}}
        for sid in sessions:
            meta = sessions_meta.get(
                sid, {"name": "Chat", "created": time.time(), "updated": time.time()}
            )
            data["sessions"][sid] = {
                "name": meta["name"],
                "created": meta["created"],
                "updated": meta["updated"],
                "messages": sessions[sid],
            }
    with open(SESSIONS_FILE, "w") as f:
        json.dump(data, f, indent=2)


def estimate_tokens(messages):
    total_chars = 0
    for msg in messages:
        content = msg.get("content", "")
        if isinstance(content, str):
            total_chars += len(content)
        elif isinstance(content, list):
            for part in content:
                if isinstance(part, dict) and part.get("type") == "text":
                    total_chars += len(part.get("text", ""))
    return max(1, total_chars // 4)


def trim_messages_for_context(messages):
    trimmed = list(messages)
    while estimate_tokens(trimmed) > MAX_INPUT_TOKENS and len(trimmed) > 1:
        trimmed.pop(0)
    return trimmed


def set_status(task_id, message):
    with _data_lock:
        if task_id in tasks:
            tasks[task_id]["status"] = "working"
            tasks[task_id]["message"] = message


def is_llama_alive():
    try:
        r = requests.get(f"{LLAMA_BASE}/health", timeout=5)
        return r.status_code == 200
    except:
        return False


_model_transition_lock = threading.Lock()


def unload_llama_model():
    global model_status
    with _model_transition_lock:
        with _data_lock:
            if model_status == "unloaded":
                return True
            model_status = "unloading"

        print("[llama] Requesting model unload from VRAM...")
        try:
            r = requests.post(
                f"{LLAMA_BASE}/models/unload", json={"model": MODEL_ID}, timeout=30
            )
            if r.status_code == 200:
                print("[llama] Model unloaded")
                with _data_lock:
                    model_status = "unloaded"
                return True
            print(f"[llama] Unload response: {r.status_code} {r.text[:200]}")
        except Exception as e:
            print(f"[llama] Unload error: {e}")

        # Check real status if unload failed or erred out
        with _data_lock:
            model_status = "chat_loaded" if is_llama_alive() else "unloaded"
        return False


def load_llama_model():
    global model_status
    with _data_lock:
        model_status = "loading"
    print(f"[llama] Sending load request for model '{MODEL_ID}'...")
    try:
        r = requests.post(
            f"{LLAMA_BASE}/models/load", json={"model": MODEL_ID}, timeout=180
        )
        if r.status_code in (200, 201):
            for i in range(30):
                if is_llama_alive():
                    print(f"[llama] Model ready (attempt {i+1})")
                    with _data_lock:
                        model_status = "chat_loaded"
                    return True
                time.sleep(2)
        else:
            print(f"[llama] Load failed ({r.status_code}): {r.text[:200]}")
    except Exception as e:
        print(f"[llama] Load exception: {e}")

    # Fallback check: verify if the server is alive and responding anyway
    if is_llama_alive():
        with _data_lock:
            model_status = "chat_loaded"
        return True

    with _data_lock:
        model_status = "unloaded"
    return False


def free_comfyui_vram():
    print("[comfyui] Freeing VRAM...")
    try:
        r = requests.post(
            f"{COMFYUI_URL}/free",
            json={"unload_models": True, "free_memory": True},
            timeout=30,
        )
        if r.status_code == 200:
            print("[comfyui] VRAM freed")
            return True
    except Exception as e:
        print(f"[comfyui] Free error: {e}")
    return False


def get_gpu_temp():
    try:
        r = subprocess.run(
            ["nvidia-smi", "--query-gpu=temperature.gpu", "--format=csv,noheader"],
            capture_output=True, text=True, timeout=5
        )
        return int(r.stdout.strip())
    except Exception:
        return None


def get_ram_usage():
    try:
        r = subprocess.run(["free", "-m"], capture_output=True, text=True, timeout=5)
        lines = r.stdout.strip().split("\n")
        parts = lines[1].split()
        total = int(parts[1])
        available = int(parts[6])
        return (total - available) / total * 100
    except Exception:
        return None


def kill_llama_server():
    subprocess.run(["pkill", "-f", "llama-server"], capture_output=True)
    time.sleep(1)
    subprocess.run(["pkill", "-9", "-f", "llama-server"], capture_output=True)


def kill_comfyui():
    subprocess.run(["pkill", "-f", "main.py.*lowvram"], capture_output=True)


def restart_servers():
    print("Restarting servers")
    kill_llama_server()
    kill_comfyui()
    time.sleep(1)
    log_dir = os.path.expanduser("~/local-ai")
    llm_log = open(os.path.join(log_dir, "llama-server.log"), "a")
    comfy_log = open(os.path.join(log_dir, "comfyui.log"), "a")
    subprocess.Popen([LLAMA_SERVER_PATH] + LLAMA_SERVER_ARGS, stdout=llm_log, stderr=llm_log, start_new_session=True)
    subprocess.Popen([os.path.join(COMFYUI_DIR, "venv/bin/python"), "main.py", "--output-directory", COMFYUI_OUTPUT, "--lowvram"], cwd=COMFYUI_DIR, stdout=comfy_log, stderr=comfy_log, start_new_session=True)
    deadline = time.time() + 120
    while time.time() < deadline:
        time.sleep(2)
        try:
            r = requests.get(f"{LLAMA_BASE}/health", timeout=3)
            if r.status_code == 200:
                print("[restart] llama-server healthy")
                return
        except Exception:
            pass
    print("[restart] llama-server did not respond within 2 minutes — killing")
    kill_llama_server()


def location_str():
    return "Unknown"


def web_search(query):
    loc_str = location_str()
    ctx = f"[Date: {datetime.now().strftime('%Y-%m-%d %A')}] [Location: {loc_str}] {query}"
    from urllib.parse import urlencode

    params = {"q": ctx, "format": "json"}
    search_url = f"{SEARXNG_URL}?{urlencode(params)}"
    print("Performing web search", search_url)
    r = requests.get(SEARXNG_URL, params=params, timeout=10)
    data = r.json()
    results = data.get("results", [])[:5]
    formatted = []
    for x in results:
        formatted.append(
            {
                "title": x.get("title", ""),
                "url": x.get("url", ""),
                "snippet": x.get("content", "") or x.get("snippet", ""),
            }
        )
    return json.dumps(
        {
            "results": formatted,
            "search_date": datetime.now().strftime("%Y-%m-%d %A"),
            "query": query,
            "search_url": search_url,
        }
    )


def generate_image(prompt, task_id, negative_prompt="", model="sd3_5_medium"):
    global model_status
    print(f"\n[image] Generating image for task {task_id} with the prompt: {prompt}")
    set_status(task_id, "Freeing VRAM for image generation...")
    unload_llama_model()

    gen_tag = str(uuid.uuid4())[:8]
    prefix = f"gen_{gen_tag}_"
    cfg = IMAGE_MODELS.get(model, IMAGE_MODELS["sd3_5_medium"])

    if model == "sd3_5_medium":
        print("Chose SD 3.5 for image generation")
        workflow = {
            "1": {"class_type": "UnetLoaderGGUF", "inputs": {"unet_name": cfg["unet"]}},
            "2": {
                "class_type": "TripleCLIPLoaderGGUF",
                "inputs": {
                    "clip_name1": cfg["clip1"],
                    "clip_name2": cfg["clip2"],
                    "clip_name3": cfg["t5"],
                    "type": "sd3",
                },
            },
            "3": {
                "class_type": "CLIPTextEncode",
                "inputs": {"text": prompt, "clip": ["2", 0]},
            },
            "4": {
                "class_type": "CLIPTextEncode",
                "inputs": {"text": negative_prompt, "clip": ["2", 0]},
            },
            "5": {
                "class_type": "EmptySD3LatentImage",
                "inputs": {"width": 512, "height": 512, "batch_size": 1},
            },
            "6": {
                "class_type": "KSampler",
                "inputs": {
                    "seed": random.randint(0, 2**31),
                    "steps": 20,  # Recommended steps for SD 3.5 Medium
                    "cfg": 4.5,  # Recommended CFG range for SD 3.5 Medium: 3.5 to 5.0
                    "sampler_name": "euler",
                    "scheduler": "sgm_uniform",
                    "denoise": 1.0,
                    "model": ["1", 0],
                    "positive": ["3", 0],
                    "negative": ["4", 0],
                    "latent_image": ["5", 0],
                },
            },
            "7": {"class_type": "VAELoader", "inputs": {"vae_name": cfg["vae"]}},
            "8": {
                "class_type": "VAEDecode",
                "inputs": {"samples": ["6", 0], "vae": ["7", 0]},
            },
            "9": {
                "class_type": "SaveImage",
                "inputs": {"filename_prefix": prefix, "images": ["8", 0]},
            },
        }
    else:
        workflow = {
            "3": {
                "class_type": "KSampler",
                "inputs": {
                    "seed": random.randint(0, 2**31),
                    "steps": 20,
                    "cfg": 7,
                    "sampler_name": "euler",
                    "scheduler": "normal",
                    "denoise": 1,
                    "model": ["4", 0],
                    "positive": ["6", 0],
                    "negative": ["7", 0],
                    "latent_image": ["5", 0],
                },
            },
            "4": {
                "class_type": "CheckpointLoaderSimple",
                "inputs": {"ckpt_name": cfg["ckpt"]},
            },
            "5": {
                "class_type": "EmptyLatentImage",
                "inputs": {"width": 512, "height": 512, "batch_size": 1},
            },
            "6": {
                "class_type": "CLIPTextEncode",
                "inputs": {"text": prompt, "clip": ["4", 1]},
            },
            "7": {
                "class_type": "CLIPTextEncode",
                "inputs": {"text": negative_prompt, "clip": ["4", 1]},
            },
            "8": {
                "class_type": "VAEDecode",
                "inputs": {"samples": ["3", 0], "vae": ["10", 0]},
            },
            "9": {
                "class_type": "SaveImage",
                "inputs": {"filename_prefix": prefix, "images": ["8", 0]},
            },
            "10": {"class_type": "VAELoader", "inputs": {"vae_name": cfg["vae"]}},
        }

    with _data_lock:
        model_status = "image_active"
        tasks[task_id]["gen_prompt"] = prompt
        tasks[task_id]["_image_model"] = model
        tasks[task_id]["negative_prompt"] = negative_prompt
    p_short = prompt[:200] + ("..." if len(prompt) > 200 else "")
    set_status(task_id, f"Generating image with ComfyUI... Prompt: {p_short}")
    try:
        r = requests.post(
            f"{COMFYUI_URL}/prompt", json={"prompt": workflow}, timeout=120
        )
        data = r.json()

        if "error" in data:
            result = json.dumps({"error": f"ComfyUI: {data['error']}"})
        else:
            prompt_id = data["prompt_id"]
            found_file = None
            for _ in range(120):
                time.sleep(1)
                try:
                    hr = requests.get(f"{COMFYUI_URL}/history/{prompt_id}", timeout=10)
                    hist = hr.json()

                    if prompt_id in hist:
                        outputs = hist[prompt_id].get("outputs", {})
                        for node_id, node_out in outputs.items():
                            for img in node_out.get("images", []):
                                fname = img["filename"]
                                fpath = os.path.join(IMG_PATH, fname)
                                found_file = fpath
                                break
                        if found_file:
                            break
                except Exception:
                    pass
            if found_file:
                tasks[task_id]["image_file"] = found_file
                set_status(task_id, f"Image saved as {found_file}")
                result = json.dumps({"prompt_id": prompt_id, "file": found_file})
            else:
                result = json.dumps({"error": "Image generation timeout"})
    except Exception as e:
        result = json.dumps({"error": str(e)})
    finally:
        set_status(task_id, "Freeing image generation VRAM...")
        free_comfyui_vram()
        set_status(task_id, "Loading chat model...")
        load_llama_model()
    return result


def edit_image(
    prompt, task_id, image_b64, negative_prompt="", denoise=0.6, model="sd3_5_medium", sid=None
):
    if not image_b64 and sid:
        with _data_lock:
            msgs = list(sessions.get(sid, []))
        print(f"[edit_image] Scanning {len(msgs)} session messages for _image_url")
        for msg in reversed(msgs):
            url = (msg.get("_image_url") or "").strip()
            role = msg.get("role", "?")
            print(f"[edit_image]  msg role={role} _image_url={url}")
            if url:
                parts = url.split("/")
                fname = parts[-1] if parts else ""
                fpath = os.path.join(IMG_PATH, fname)
                print(f"[edit_image]  checking fpath={fpath} exists={os.path.exists(fpath)}")
                if os.path.exists(fpath):
                    with open(fpath, "rb") as f:
                        image_b64 = base64.b64encode(f.read()).decode()
                    break

    if not image_b64:
        print("[edit_image] FAILED to find an image to edit")
        return json.dumps({"error": "No image provided for editing."})
    print(f"[edit_image] Found image ({len(image_b64)} bytes base64), proceeding with edit")

    print(f"\n[image_edit] Editing image for task {task_id} with prompt: {prompt}")
    set_status(task_id, "Freeing VRAM for image editing...")
    unload_llama_model()

    gen_tag = str(uuid.uuid4())[:8]
    prefix = f"edit_{gen_tag}_"
    input_filename = f"input_{gen_tag}.png"

    try:
        import folder_paths
        input_dir = folder_paths.get_input_directory()
    except Exception:
        input_dir = os.path.join(COMFYUI_DIR, "input")
    os.makedirs(input_dir, exist_ok=True)
    input_filepath = os.path.join(input_dir, input_filename)

    with open(input_filepath, "wb") as f:
        f.write(base64.b64decode(image_b64))

    cfg = IMAGE_MODELS.get(model, IMAGE_MODELS["sd3_5_medium"])

    # SD 3.5 Medium Img2Img Workflow
    workflow = {
        "1": {"class_type": "UnetLoaderGGUF", "inputs": {"unet_name": cfg["unet"]}},
        "2": {
            "class_type": "TripleCLIPLoaderGGUF",
            "inputs": {
                "clip_name1": cfg["clip1"],
                "clip_name2": cfg["clip2"],
                "clip_name3": cfg["t5"],
                "type": "sd3",
            },
        },
        "3": {
            "class_type": "CLIPTextEncode",
            "inputs": {"text": prompt, "clip": ["2", 0]},
        },
        "4": {
            "class_type": "CLIPTextEncode",
            "inputs": {"text": negative_prompt, "clip": ["2", 0]},
        },
        "5": {
            "class_type": "LoadImage",
            "inputs": {"image": input_filename},
        },
        "7": {"class_type": "VAELoader", "inputs": {"vae_name": cfg["vae"]}},
        "5_encode": {
            "class_type": "VAEEncode",
            "inputs": {
                "pixels": ["5", 0],
                "vae": ["7", 0],
            },
        },
        "6": {
            "class_type": "KSampler",
            "inputs": {
                "seed": random.randint(0, 2**31),
                "steps": 20,
                "cfg": 4.5,
                "sampler_name": "euler",
                "scheduler": "sgm_uniform",
                "denoise": float(denoise),
                "model": ["1", 0],
                "positive": ["3", 0],
                "negative": ["4", 0],
                "latent_image": ["5_encode", 0],
            },
        },
        "8": {
            "class_type": "VAEDecode",
            "inputs": {"samples": ["6", 0], "vae": ["7", 0]},
        },
        "9": {
            "class_type": "SaveImage",
            "inputs": {"filename_prefix": prefix, "images": ["8", 0]},
        },
    }

    with _data_lock:
        model_status = "image_active"
        tasks[task_id]["gen_prompt"] = prompt
        tasks[task_id]["_image_model"] = model
        tasks[task_id]["negative_prompt"] = negative_prompt

    set_status(task_id, f"Editing image with ComfyUI... Prompt: {prompt[:150]}")

    try:
        r = requests.post(
            f"{COMFYUI_URL}/prompt", json={"prompt": workflow}, timeout=120
        )
        data = r.json()

        if "error" in data:
            result = json.dumps({"error": f"ComfyUI: {data['error']}"})
        else:
            prompt_id = data["prompt_id"]
            found_file = None
            for _ in range(120):
                time.sleep(1)
                try:
                    hr = requests.get(f"{COMFYUI_URL}/history/{prompt_id}", timeout=10)
                    hist = hr.json()
                    if prompt_id in hist:
                        outputs = hist[prompt_id].get("outputs", {})
                        for node_id, node_out in outputs.items():
                            for img in node_out.get("images", []):
                                fname = img["filename"]
                                found_file = os.path.join(IMG_PATH, fname)
                                break
                        if found_file:
                            break
                except Exception:
                    pass

            if found_file:
                tasks[task_id]["image_file"] = found_file
                set_status(task_id, f"Edited image saved as {found_file}")
                result = json.dumps({"prompt_id": prompt_id, "file": found_file})
            else:
                result = json.dumps({"error": "Image editing timeout"})
    except Exception as e:
        result = json.dumps({"error": str(e)})
    finally:
        set_status(task_id, "Freeing image generation VRAM...")
        free_comfyui_vram()
        set_status(task_id, "Loading chat model...")
        load_llama_model()

    return result


def _event_post(ev_type, task_id, **data):
    _event_queue.put((ev_type, task_id, data))


def _llm_worker(task_id, sid, payload, round_num):
    try:
        r = requests.post(LLAMA_URL, json=payload, timeout=600)
        body = r.json()
        if "choices" in body:
            _event_post("llm_ok", task_id, body=body, round=round_num, sid=sid)
        else:
            _event_post("llm_err", task_id, error=f"Unexpected response ({r.status_code}): {str(body)[:300]}", round=round_num, sid=sid)
    except Exception as e:
        _event_post("llm_err", task_id, error=str(e), round=round_num, sid=sid)


def _tool_worker(task_id, sid, tc, image_b64, round_num, tool_index):
    tool_name = tc["function"]["name"]
    try:
        args = json.loads(tc["function"]["arguments"])
    except Exception:
        args = {}

    with _data_lock:
        tu = list(tasks.get(task_id, {}).get("_tools_used", []))
    has_generated_image = "generate_image" in tu

    if tool_name == "web_search":
        set_status(task_id, f"Searching web for: {args.get('query')}...")
        result = web_search(args["query"])
        with _data_lock:
            t = tasks.get(task_id)
            if t:
                t.setdefault("_tools_used", []).append(tool_name)
                try:
                    t.setdefault("_search_details", []).append(json.loads(result))
                except Exception:
                    pass
        _event_post("tool_ok", task_id, tc_id=tc["id"], result=result, sid=sid, round=round_num, tool_index=tool_index)

    elif tool_name == "edit_image":
        result = edit_image(
            prompt=args.get("prompt", ""),
            task_id=task_id,
            image_b64=image_b64,
            negative_prompt=args.get("negative_prompt", ""),
            denoise=args.get("denoise", 0.6),
            model="sd3_5_medium",
            sid=sid,
        )
        res_data = json.loads(result)
        if "file" in res_data:
            image_url = f"/output/{os.path.basename(res_data['file'])}"
            with _data_lock:
                t = tasks.get(task_id)
                if t:
                    t.setdefault("_tools_used", []).append(tool_name)
            msg_entry = {
                "role": "assistant",
                "content": "Here is your generated image:",
                "_tools_used": tu + [tool_name],
                "_image_url": image_url,
                "_gen_prompt": args.get("prompt", ""),
                "_image_model": None,
            }
            with _data_lock:
                if sid in sessions:
                    sessions[sid].append(msg_entry)
                    sessions_meta.setdefault(sid, {})["updated"] = time.time()
            save_sessions()
            _event_post("img_done", task_id, image_url=image_url, tools_used=tu + [tool_name], gen_prompt=args.get("prompt", ""), image_model=None, sid=sid)
        else:
            _event_post("tool_ok", task_id, tc_id=tc["id"], result=result, sid=sid, round=round_num, tool_index=tool_index)

    elif tool_name == "generate_image":
        if has_generated_image:
            result = json.dumps({"error": "Image generation limit reached for this prompt."})
            _event_post("tool_ok", task_id, tc_id=tc["id"], result=result, sid=sid, round=round_num, tool_index=tool_index)
        else:
            result = generate_image(
                prompt=args.get("prompt", ""),
                task_id=task_id,
                negative_prompt=args.get("negative_prompt", ""),
                model=args.get("model") or "sd3_5_medium",
            )
            res_data = json.loads(result)
            if "file" in res_data:
                image_url = f"/output/{os.path.basename(res_data['file'])}"
                image_model_s = args.get("model") or "sd3_5_medium"
                with _data_lock:
                    t = tasks.get(task_id)
                    if t:
                        t.setdefault("_tools_used", []).append(tool_name)
                msg_entry = {
                    "role": "assistant",
                    "content": "Here is your generated image:",
                    "_tools_used": tu + [tool_name],
                    "_image_url": image_url,
                    "_gen_prompt": args.get("prompt", ""),
                    "_image_model": image_model_s,
                }
                with _data_lock:
                    if sid in sessions:
                        sessions[sid].append(msg_entry)
                        sessions_meta.setdefault(sid, {})["updated"] = time.time()
                save_sessions()
                _event_post("img_done", task_id, image_url=image_url, tools_used=tu + [tool_name], gen_prompt=args.get("prompt", ""), image_model=image_model_s, sid=sid)
            else:
                _event_post("tool_ok", task_id, tc_id=tc["id"], result=result, sid=sid, round=round_num, tool_index=tool_index)
    else:
        result = json.dumps({"error": f"Unknown tool: {tool_name}"})
        _event_post("tool_ok", task_id, tc_id=tc["id"], result=result, sid=sid, round=round_num, tool_index=tool_index)


def _prepare_session(task_id, sid, user_message, image_b64):
    date_loc_context = f"[Current date: {datetime.now().strftime('%Y-%m-%d %A %H:%M')}] [User location: {location_str()}]"
    full_sys_content = f"{SYS_CONTENT}\n\n{date_loc_context}"
    with _data_lock:
        if sid not in sessions or not sessions[sid]:
            sessions[sid] = [{"role": "system", "content": full_sys_content}]
        elif sessions[sid][0].get("role") == "system":
            sessions[sid][0]["content"] = full_sys_content
        else:
            sessions[sid].insert(0, {"role": "system", "content": full_sys_content})
        if sid not in sessions_meta:
            sessions_meta[sid] = {"name": user_message[:50], "created": time.time(), "updated": time.time()}
        content = []
        if image_b64:
            content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"}})
        content.append({"type": "text", "text": user_message})
        sessions[sid].append({"role": "user", "content": content})
        if sessions_meta[sid]["name"] in ("New Chat", ""):
            sessions_meta[sid]["name"] = user_message[:50] + ("..." if len(user_message) > 50 else "")
        sessions_meta[sid]["updated"] = time.time()
    save_sessions()
    with _data_lock:
        ms = model_status
    if ms != "chat_loaded":
        load_llama_model()


def _start_llm_round(task_id, sid, round_num):
    with _data_lock:
        t = tasks.get(task_id)
        if not t:
            return
        t["_state"] = "llm_waiting"
        t["_round"] = round_num
        messages = trim_messages_for_context(sessions.get(sid, []))
    payload = {"model": MODEL_ID, "messages": messages, "tools": TOOLS, "tool_choice": "auto", "max_tokens": 4096}
    set_status(task_id, "Thinking..." if round_num == 0 else f"Thinking (round {round_num})...")
    _llm_pool.submit(_llm_worker, task_id, sid, payload, round_num)


def _finalize_task(task_id, sid, msg_content, body):
    with _data_lock:
        t = tasks.get(task_id)
        if not t:
            return
        tools_used = list(t.get("_tools_used", []))
        search_details = list(t.get("_search_details", []))
        image_filename = t.get("image_file")
        gen_prompt = t.get("gen_prompt")
        image_model = t.get("_image_model")
    image_url = f"/output/{image_filename}" if image_filename else None
    timings = body.get("timings", {})
    predicted_per_second = timings.get("predicted_per_second")
    msg_entry = {
        "role": "assistant", "content": msg_content,
        "_tools_used": tools_used, "_image_url": image_url,
        "_gen_prompt": gen_prompt, "_image_model": image_model,
        "_search_details": search_details,
    }
    with _data_lock:
        if sid in sessions:
            sessions[sid].append(msg_entry)
            sessions_meta.setdefault(sid, {})["updated"] = time.time()
        _last_tps = predicted_per_second
        _last_llm_use = time.time()
    save_sessions()
    with _data_lock:
        if task_id in tasks:
            tasks[task_id] = {
                "status": "done", "response": msg_content, "session_id": sid,
                "token_estimate": estimate_tokens(sessions.get(sid, [])),
                "predicted_per_second": predicted_per_second,
                "tools_used": tools_used, "image": image_url, "_image_url": image_url,
                "gen_prompt": gen_prompt, "_image_model": image_model,
                "_search_details": search_details,
            }


def _set_task_error(task_id, error, sid=None):
    with _data_lock:
        if task_id in tasks:
            d = tasks[task_id]
            tasks[task_id] = {"status": "error", "error": str(error), "session_id": d.get("session_id", sid)}


def _event_loop():
    global _current_task_id
    while True:
        ev_type, task_id, data = _event_queue.get()
        t = tasks.get(task_id)
        if not t:
            continue

        if ev_type == "start":
            sid = data["sid"]
            user_message = data["message"]
            image_b64 = data.get("image")
            with _data_lock:
                tasks[task_id] = {"status": "working", "message": "Processing task...", "session_id": sid, "_tools_used": [], "_search_details": [], "_original_message": user_message, "_original_image": image_b64}
            _current_task_id = task_id
            _prepare_session(task_id, sid, user_message, image_b64)
            _start_llm_round(task_id, sid, 0)

        elif ev_type == "llm_ok":
            if t.get("_state") != "llm_waiting":
                continue
            sid = data["sid"]
            round_num = data["round"]
            body = data["body"]
            msg = body["choices"][0]["message"]
            with _data_lock:
                _last_llm_use = time.time()
            if msg.get("tool_calls"):
                with _data_lock:
                    tt = tasks.get(task_id)
                    if tt:
                        tt.setdefault("_tools_used", [])
                        tt.setdefault("_search_details", [])
                pending = len(msg["tool_calls"])
                with _data_lock:
                    tt = tasks.get(task_id)
                    if tt:
                        tt["_state"] = "tools_running"
                        tt["_pending_tools"] = pending
                for i, tc in enumerate(msg["tool_calls"]):
                    _tool_pool.submit(_tool_worker, task_id, sid, tc, t.get("_image_b64"), round_num, i)
            else:
                _finalize_task(task_id, sid, msg.get("content", ""), body)

        elif ev_type == "llm_err":
            if t.get("_state") != "llm_waiting":
                continue
            _set_task_error(task_id, data["error"], data.get("sid"))

        elif ev_type == "tool_ok":
            sid = data["sid"]
            tc_id = data["tc_id"]
            result = data["result"]
            with _data_lock:
                if sid in sessions:
                    sessions[sid].append({"role": "tool", "tool_call_id": tc_id, "content": result})
                    sessions_meta.setdefault(sid, {})["updated"] = time.time()
                tt = tasks.get(task_id)
                if not tt or tt.get("status") in ("done", "error"):
                    continue
                pending = (tt.get("_pending_tools", 0) - 1) if tt else 0
                if tt:
                    tt["_pending_tools"] = pending
            save_sessions()
            if pending <= 0:
                round_num = data.get("round", 0) + 1
                with _data_lock:
                    tt = tasks.get(task_id)
                    if tt:
                        tt["_round"] = round_num
                if round_num < 10:
                    _start_llm_round(task_id, sid, round_num)
                else:
                    _set_task_error(task_id, "Max tool rounds exceeded", sid)

        elif ev_type == "tool_err":
            result = data.get("result", json.dumps({"error": data.get("error", "Tool error")}))
            with _data_lock:
                if data.get("sid") in sessions:
                    sessions[data["sid"]].append({"role": "tool", "tool_call_id": data["tc_id"], "content": result})
                    sessions_meta.setdefault(data["sid"], {})["updated"] = time.time()
                tt = tasks.get(task_id)
                if not tt or tt.get("status") in ("done", "error"):
                    continue
                pending = (tt.get("_pending_tools", 0) - 1) if tt else 0
                if tt:
                    tt["_pending_tools"] = pending
            save_sessions()
            if pending <= 0:
                round_num = data.get("round", 0) + 1
                with _data_lock:
                    tt = tasks.get(task_id)
                    if tt:
                        tt["_round"] = round_num
                if round_num < 10:
                    _start_llm_round(task_id, data["sid"], round_num)
                else:
                    _set_task_error(task_id, "Max tool rounds exceeded", data["sid"])

        elif ev_type == "img_done":
            image_url = data["image_url"]
            sid = data["sid"]
            tools_used = data["tools_used"]
            gen_prompt = data["gen_prompt"]
            image_model = data.get("image_model")
            with _data_lock:
                if task_id in tasks:
                    tasks[task_id] = {
                        "status": "done", "response": "Here is your generated image:",
                        "session_id": sid, "image": image_url, "_image_url": image_url,
                        "tools_used": tools_used, "gen_prompt": gen_prompt,
                        "_image_model": image_model,
                    }


def _queue_worker():
    global _current_task_id
    while True:
        item = None
        with _queue_lock:
            while not _task_queue:
                _queue_cond.wait()
            with _data_lock:
                oh = _overheated
            if oh or _ram_evacuating:
                label = "GPU overheating" if oh else "RAM pressure — restarting servers"
                for qitem in _task_queue:
                    tid = qitem["task_id"]
                    if tid in tasks:
                        tasks[tid] = {"status": "waiting", "message": f"Server paused — {label}. Will resume shortly.", "session_id": qitem["session_id"]}
                _queue_cond.wait(5)
                continue
            item = _task_queue.pop(0)
            _current_task_id = item["task_id"]
        _event_post("start", item["task_id"], sid=item["session_id"], message=item["message"], image=item.get("image"))
        # Wait for this task to finish (status becomes "done" or "error") before dequeuing the next
        while True:
            with _data_lock:
                st = tasks.get(item["task_id"], {}).get("status")
            if st in ("done", "error"):
                break
            time.sleep(0.5)
        with _queue_lock:
            _current_task_id = None
            _queue_cond.notify_all()


def _idle_unload_loop():
    global _last_llm_use
    while True:
        time.sleep(10)

        with _queue_lock:
            queue_active = len(_task_queue) > 0 or _current_task_id is not None

        with _data_lock:
            ms = model_status
            lu = _last_llm_use

        # Only unload if loaded, inactive for > 300s, and no queue tasks pending
        if ms == "chat_loaded" and (time.time() - lu > 300) and not queue_active:
            print("[idle] No LLM activity for 300s, releasing VRAM model weights...")
            unload_llama_model()


def _evacuate_ram():
    global _current_task_id, _ram_evacuating
    _ram_evacuating = True
    print("[ram] Emergency RAM evacuation")
    with _queue_lock:
        tid = _current_task_id
        if tid:
            with _data_lock:
                t = tasks.get(tid)
                if t and t.get("status") not in ("done", "error"):
                    entry = {
                        "task_id": tid,
                        "session_id": t.get("session_id", ""),
                        "message": t.get("_original_message", ""),
                        "image": t.get("_original_image"),
                    }
                    _task_queue.insert(0, entry)
                    t["status"] = "error"
                    t["error"] = "Server ran out of RAM — requeued"
                    t["_ram_evacuating"] = True
                    print(f"[ram] Requeued task {tid} to front of queue")
    kill_llama_server()
    kill_comfyui()
    print("[ram] Killed llama-server and ComfyUI")
    while True:
        time.sleep(5)
        ram = get_ram_usage()
        if ram is not None and ram <= RAM_RESUME_THRESHOLD:
            print(f"[ram] RAM {ram:.0f}% ≤ {RAM_RESUME_THRESHOLD}%, restarting servers")
            break
    restart_servers()
    _ram_evacuating = False


def _thermal_monitor():
    global _overheated, _gpu_temp
    while True:
        time.sleep(10)
        temp = get_gpu_temp()
        with _data_lock:
            _gpu_temp = temp
            if temp is not None and temp >= TEMP_THRESHOLD_ON:
                if not _overheated:
                    print(f"[thermal] GPU {temp}°C >= {TEMP_THRESHOLD_ON}°C, OVERHEATED")
                    _overheated = True
            elif _overheated and (temp is None or temp <= TEMP_THRESHOLD_OFF):
                print(f"[thermal] GPU {temp}°C <= {TEMP_THRESHOLD_OFF}°C, resumed")
                _overheated = False

        if _overheated:
            with _queue_lock:
                busy = _current_task_id is not None
            if not busy:
                with _data_lock:
                    ms = model_status
                if ms == "chat_loaded":
                    print("[thermal] Overheated — unloading chat model")
                    unload_llama_model()
                elif ms == "image_active":
                    print("[thermal] Overheated — freeing ComfyUI VRAM")
                    free_comfyui_vram()

        if not _ram_evacuating:
            ram = get_ram_usage()
            if ram is not None and ram >= RAM_EVAC_THRESHOLD:
                print(f"[ram] RAM usage {ram:.0f}% >= {RAM_EVAC_THRESHOLD}%")
                _evacuate_ram()


HTML_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")
try:
    with open(HTML_FILE) as f:
        HTML = f.read()
except:
    HTML = "<html><body><h1>index.html missing</h1></body></html>"


def extract_file_text(name, data_b64):
    ext = os.path.splitext(name)[1].lower()
    raw = base64.b64decode(data_b64)
    if ext == ".pdf":
        try:
            import fitz

            doc = fitz.open(stream=raw, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            return text
        except ImportError:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(raw)
                tmp = f.name
            try:
                r = subprocess.run(
                    ["pdftotext", tmp, "-"], capture_output=True, text=True, timeout=30
                )
                return r.stdout
            finally:
                os.unlink(tmp)
    elif ext == ".docx":
        from docx import Document

        doc = Document(io.BytesIO(raw))
        return "\n".join(p.text for p in doc.paragraphs)
    elif ext == ".doc":
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as f:
            f.write(raw)
            tmp = f.name
        try:
            r = subprocess.run(
                ["catdoc", tmp], capture_output=True, text=True, timeout=30
            )
            if r.returncode == 0:
                return r.stdout
            r = subprocess.run(
                ["antiword", tmp], capture_output=True, text=True, timeout=30
            )
            return r.stdout
        finally:
            os.unlink(tmp)
    elif ext in (".xls", ".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(rows)
    return ""


class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/api/model-status":
            with _data_lock:
                ms, tps, oh, gtemp = model_status, _last_tps, _overheated, _gpu_temp
            self.send_json({"model": ms, "predicted_per_second": tps, "overheated": oh, "gpu_temp": gtemp})
        elif self.path.startswith("/output/"):
            filename = os.path.basename(urlparse(self.path).path)
            fpath = os.path.abspath(os.path.join(COMFYUI_OUTPUT, filename))
            if fpath.startswith(os.path.abspath(COMFYUI_OUTPUT)) and os.path.exists(
                fpath
            ):
                self.send_response(200)
                self.send_header("Content-Type", "image/png")
                self.end_headers()
                with open(fpath, "rb") as f:
                    self.wfile.write(f.read())
                return
            self.send_error(404)
        elif self.path.startswith("/api/status/"):
            task_id = os.path.basename(self.path)
            with _data_lock:
                status = tasks.get(
                    task_id, {"status": "unknown", "message": "Not found"}
                )
            self.send_json(status)
        elif self.path == "/api/sessions":
            with _data_lock:
                sorted_items = sorted(
                    sessions_meta.items(),
                    key=lambda x: x[1].get("updated", 0),
                    reverse=True,
                )
                result = [
                    {
                        "session_id": sid,
                        "name": meta.get("name", "Chat"),
                        "created": meta.get("created", 0),
                        "updated": meta.get("updated", 0),
                        "token_estimate": estimate_tokens(sessions.get(sid, [])),
                    }
                    for sid, meta in sorted_items
                ]
            self.send_json(result)
        elif self.path.startswith("/api/sessions/") and self.path.endswith("/messages"):
            sid = self.path.split("/")[3]
            with _data_lock:
                msgs = sessions.get(sid)
            if msgs is not None:
                self.send_json(
                    {
                        "messages": msgs,
                        "token_estimate": estimate_tokens(msgs),
                    }
                )
            else:
                self.send_error(404)
        elif self.path == "/":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(HTML.encode())
        else:
            self.send_error(404)

    def do_DELETE(self):
        if self.path.startswith("/api/sessions/"):
            sid = self.path.split("/")[3]
            with _data_lock:
                msgs = list(sessions.get(sid, []))
            for msg in msgs:
                if msg.get("role") == "assistant":
                    url = msg.get("_image_url", "") or ""
                    if url:
                        fname = os.path.basename(url)
                        fpath = os.path.join(IMG_PATH, fname)
                        if os.path.exists(fpath):
                            os.remove(fpath)
            with _data_lock:
                exists = sid in sessions
                if exists:
                    sessions.pop(sid, None)
                    sessions_meta.pop(sid, None)
            if exists:
                save_sessions()
                self.send_json({"status": "deleted"})
            else:
                self.send_error(404)
        else:
            self.send_error(404)

    def do_PUT(self):
        if self.path.startswith("/api/sessions/"):
            sid = self.path.split("/")[3]
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
            with _data_lock:
                meta = sessions_meta.get(sid)
                if meta:
                    meta["name"] = body.get("name", meta["name"])
                    meta["updated"] = time.time()
            if meta:
                save_sessions()
                self.send_json({"status": "updated"})
            else:
                self.send_error(404)
        else:
            self.send_error(404)

    def do_POST(self):
        if self.path == "/api/chat":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
            task_id = str(uuid.uuid4())
            sid = body.get("session_id", "default")
            with _data_lock:
                if _overheated:
                    self.send_json({"error": "Server overloaded — your message is queued and will be processed once the GPU cools down"}, status=503)
                    return

            entry = {
                "task_id": task_id,
                "session_id": sid,
                "message": body.get("message", ""),
                "image": body.get("image"),
            }
            with _queue_lock:
                if len(_task_queue) >= MAX_QUEUE_SIZE:
                    self.send_json({"error": "Server busy"}, status=503)
                    return
                _task_queue.append(entry)
                _queue_cond.notify()
            with _data_lock:
                tasks[task_id] = {
                    "status": "queued",
                    "message": "Waiting in line...",
                    "session_id": sid,
                }
            self.send_json({"task_id": task_id})
        elif self.path == "/api/extract-file":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
            text = extract_file_text(body.get("name", ""), body.get("data", ""))
            if text:
                self.send_json({"text": text})
            else:
                self.send_json(
                    {"error": "Could not extract text from file"}, status=400
                )
        elif self.path == "/api/sessions":
            sid = str(uuid.uuid4())
            now = time.time()
            with _data_lock:
                sessions[sid] = []
                sessions_meta[sid] = {
                    "name": "New Chat",
                    "created": now,
                    "updated": now,
                }
            save_sessions()
            self.send_json({"session_id": sid})
        elif self.path == "/api/location":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
            global _client_location
            _client_location = body
            self.send_json({"ok": True})
        else:
            self.send_error(404)

    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())

    def log_message(self, format, *args):
        pass


if __name__ == "__main__":
    load_sessions()
    try:
        r = requests.get(f"{LLAMA_BASE}/health", timeout=3)
        if r.status_code != 200:
            raise Exception("health check failed")
        print("[startup] llama-server is running")
    except Exception:
        print("[startup] llama-server not reachable — starting...")
        restart_servers()
    threading.Thread(target=_event_loop, daemon=True).start()
    threading.Thread(target=_queue_worker, daemon=True).start()
    threading.Thread(target=_idle_unload_loop, daemon=True).start()
    threading.Thread(target=_thermal_monitor, daemon=True).start()
    print(f"Chat UI running on http://localhost:{PORT}")
    s = http.server.HTTPServer((HOST, PORT), Handler)
    try:
        s.serve_forever()
    except KeyboardInterrupt:
        s.shutdown()
